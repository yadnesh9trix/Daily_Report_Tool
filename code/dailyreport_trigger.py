#Import libraries
import pandas as pd
import datetime
import os
import warnings
warnings.filterwarnings('ignore')
import daily_report_process as drp
import excel_writer as ew
from openpyxl import load_workbook
from datetime import timedelta

#-----------------------------------------------------------------------------------------------------------------------
# Defined todays date folder
today = datetime.datetime.today().date()
tday  = today.strftime("%d_%b_%Y")
yesterday = today - timedelta(days=0)

#-----------------------------------------------------------------------------------------------------------------------
def report_trigger(std_path,in_path,outpth,mappath,logopath,mailreport):

    # global tdsheetname, ytdsheetname
    #
    # def get_sheetnames_xlsx(filepath):
    #     wb = load_workbook(filepath, read_only=True, keep_links=False)
    #     return wb.sheetnames

    if os.path.isdir(in_path):
        files = os.listdir(in_path)
        if len(files) > 0:
            fil = files[0].lower()
            if (fil.__contains__("list")) | (fil.__contains__("amount")):
                infile = in_path + "/" + files[0]
                # sheetname = get_sheetnames_xlsx(infile)
                # for x in sheetname:
                #     if (len(x) >= 20) | (x == str(today)) | (x == 'Total') :
                #         ytdsheetname = x
                #         tdsheetname  =None
                #     # elif len(x) >= 8 | (x==str(today)):
                #     elif x == str(today):
                #         tdsheetname = x
                #     else:
                #         pass
                zonemap,usemap,construcmap,occpmap,subusemap,gatnamemap,msterdatapath_ = drp.mapping_type(mappath)
                TDCollection,tdreport = drp.zonegatwise_TDdailyreport(infile,mappath,outpth,zonemap,mailreport)

                YTD_RecoveryCollection,str_tomw  = drp.totaltax_collectionreport(std_path,infile,mappath,msterdatapath_,
                                                                                 zonemap,tdreport,mailreport)
                # YTD_RecoveryCollection = pd.DataFrame()
                if os.path.isdir(outpth):
                    print("Already Present Today's Date Folder")
                else:
                    os.mkdir(outpth)
                ew.excelwriter(outpth,logopath,TDCollection,YTD_RecoveryCollection,str_tomw)
            else:
                print("({})paid amount data not found".format(today))
        else:
            print("Data Not found")
    else:
        print("({})'s data folder not found".format(today))
#-----------------------------------------------------------------------------------------------------------------------

if __name__ == '__main__':
    std_path = r"D:\Daily_Report_Tool/"
    in_path = std_path + "Input/" + str(today) + "/"
    outpth = std_path + "Output/" + str(today) + "/"
    mappath = std_path + "Mapping/"
    logopath = std_path + "logo/"
    mailreport = std_path + "Mail_report/"

    print('Your Tool is running.\nPlease wait...\n============================='
          '===============================================================================')
    report_trigger(std_path,in_path,outpth,mappath,logopath,mailreport)

#-----------------------------------------------------------------------------------------------------------------------
